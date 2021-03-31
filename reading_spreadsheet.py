import pandas as pd
from dateutil import parser
import openpyxl
from database_configuration import DatabaseConfiguration
from database_insertion import DatabaseInsertion
import re
import MySQLdb.cursors
import os
from dotenv import load_dotenv

load_dotenv()

connection = MySQLdb.connect(
    host=os.getenv("host"), user=os.getenv("user"), passwd=os.getenv("password"), db=os.getenv("database"),
    cursorclass=MySQLdb.cursors.DictCursor
)

salt_stats_connection = MySQLdb.connect(
    host=os.getenv("salt_stats_host"), user=os.getenv("salt_stats_user"), passwd=os.getenv("salt_stats_password"),
    db=os.getenv("salt_stats_database"), cursorclass=MySQLdb.cursors.DictCursor
)


def student_project_phd_numbers(proposal_code, thesis_type_id):
    with connection.cursor() as cur:
        sql = """SELECT COUNT(*) AS phd_numbers FROM P1Thesis p1t
                 JOIN ProposalCode pc on p1t.ProposalCode_Id = pc.ProposalCode_Id
                 WHERE pc.Proposal_Code= %(proposal_code)s and ThesisType_Id = %(thesis_type_id)s
                 """
        cur.execute(sql, dict(proposal_code=proposal_code, thesis_type_id=thesis_type_id))
        results = cur.fetchall()
        return results[0]["phd_numbers"]


def student_project_msc_numbers(proposal_code, thesis_type_id):
    with connection.cursor() as cur:
        sql = """SELECT COUNT(*) AS msc_numbers FROM P1Thesis p1t
                 JOIN ProposalCode pc on p1t.ProposalCode_Id = pc.ProposalCode_Id
                 WHERE pc.Proposal_Code= %(proposal_code)s and ThesisType_Id = %(thesis_type_id)s
              """
        cur.execute(sql, dict(proposal_code=proposal_code, thesis_type_id=thesis_type_id))
        results = cur.fetchall()
        return results[0]["msc_numbers"]


def institutes(proposal_code):
    arr = []
    with connection.cursor() as cur:
        sql = """SELECT InstituteName_Name AS Institute, CONCAT(Surname,' ',FirstName ) AS ProposalInvestigator, 
             ActOnAlert as TargetOfOpportunity
             FROM  ProposalCode pc
             JOIN ProposalInvestigator pi ON pc.ProposalCode_Id = pi.ProposalCode_Id
             JOIN  Investigator ON pi.Investigator_Id = Investigator.Investigator_Id
             JOIN Institute ON Investigator.Institute_Id = Institute.Institute_Id
             JOIN InstituteName ON Institute.InstituteName_Id = InstituteName.InstituteName_Id
             JOIN ProposalGeneralInfo pg ON pc.ProposalCode_Id = pg.ProposalCode_Id
             WHERE Proposal_Code = %(proposal_code)s """
        cur.execute(sql, dict(proposal_code=proposal_code))
        results = cur.fetchall()
        for result in results:
            arr.append(result["Institute"])
        return {"institutes": ",".join(list(dict.fromkeys(arr))),
                "target of opportunity": result["TargetOfOpportunity"]}


def proposal_investigator(proposal_code):
    with connection.cursor() as cur:
        sql = """SELECT CONCAT(Surname,' ',FirstName ) AS ProposalInvestigator
                 FROM  ProposalCode pc
                 JOIN ProposalInvestigator pi ON pc.ProposalCode_Id = pi.ProposalCode_Id
                 JOIN Investigator ON pi.Investigator_Id = Investigator.Investigator_Id
                 JOIN ProposalContact ON pc.ProposalCode_Id = ProposalContact.ProposalCode_Id
                 WHERE Proposal_Code = %(proposal_code)s AND pi.Investigator_Id = Leader_Id"""
        cur.execute(sql, dict(proposal_code=proposal_code))
        results = cur.fetchall()
        for result in results:
            return result


salt_statistics_db_config = DatabaseConfiguration(
    username=os.getenv("salt_stats_user"),
    password=os.getenv("salt_stats_password"),
    host=os.getenv("salt_stats_host"),
    port=3306,
    database=os.getenv("salt_stats_database")
)
salt_stats = DatabaseInsertion(salt_statistics_db_config)


def semester_and_year_sdb(proposal_code):
    with connection.cursor() as cur:
        sql = """SELECT Semester, Year
                 FROM ProposalCode
                 JOIN Proposal ON sdb_daily.ProposalCode.ProposalCode_Id = sdb_daily.Proposal.ProposalCode_Id
                 JOIN Semester ON sdb_daily.Proposal.Semester_Id = Semester.Semester_Id
                 WHERE ProposalCode.Proposal_Code = %(proposal_code)s """
        cur.execute(sql, dict(proposal_code=proposal_code))
        results = cur.fetchall()
        return {"year": results[0]["Year"], "semester": results[0]["Semester"]}


def insert_type_of_publication():
    publication_types = ["data", "science", 'instrument', "no", "non-ref"]
    with salt_stats_connection.cursor() as cur:
        sql = """INSERT INTO PublicationType(PublicationType) VALUE (%(publication_type)s)
                 ON DUPLICATE KEY UPDATE PublicationType = %(publication_type)s 
              """
        for publication in publication_types:
            cur.execute(sql, dict(publication_type=publication))
            salt_stats_connection.commit()


insert_type_of_publication()


def insert_position_of_first_author():
    positions = ["Staff", "SA", "PhD student", "MSc student", "SALT Chairman", "Collaboration",
                 "student", "Head SALT Ops"]
    with salt_stats_connection.cursor() as cur:
        sql = """ INSERT INTO FirstAuthorPosition(AuthorPosition) VALUE(%(position)s)
                  ON DUPLICATE KEY UPDATE AuthorPosition = %(position)s """
        for position in positions:
            cur.execute(sql, dict(position=position))
            salt_stats_connection.commit()


insert_position_of_first_author()


def fixing_bad_column_return_none(column):
    """
    We check if a row contains a null value("nan/Nan" since we use pandas) or any other data which
    does not make sense and return None if it's bad data such as a
    'dash', "N/A" or empty space or return the value if it contains
    good data
    :param column: data in column of spreadsheet
    :return: None or value of column
    """
    if pd.isnull(column) or "--" in str(column):
        return None
    elif column == ' ' or column == "-" or "?" in str(column):
        return None
    elif column == "N/A":
        return None
    else:
        return column


def publication_date(year, month, day, df, ads_link):
    """
    We check if the year and month of a publication are not null("nan/Nan") and return a date for a
    publication using that year and month and the date as the 15th for every publication. Example: 2019-01-15
    :param df: pandas dataframe
    :param ads_link:ads link for publication
    :param year: Year of publication
    :param month: month of publication
    :param day: day in which publication was released (decided to use the 15th for each publication)
    :return: A publication date if the year and month are available else we return just the year of the
    publication
    """

    if not pd.isnull(year) and not pd.isnull(month) and not isinstance(month, str):
        spreadsheet_date = parser.parse("{}-{}-{}".format(year, int(month), day)).strftime("%Y-%m-%d")
    elif pd.isnull(month) and not pd.isnull(year):
        cond1 = df["Year"] == year
        cond2 = df["ADS link"] == ads_link
        row_number = df[cond1 & cond2].index[0] + 2
        raise ValueError("The Month was not added on row {}. Please add it".format(row_number))
    else:
        spreadsheet_date = None
    return spreadsheet_date


def science_types():
    all_science_types = {"exg": "Extragalactic", "sne": "Supernovae and GrW", "hz": "high z", "qso": "quasar",
                         "gal": "galaxy", "exo": "extra-solar planetary", "bin": "binary", "gc": "globular cluster",
                         "He": "Helium", "xrb": "X/gamma-ray binaries", "Gal": "Gal object", "ast": "asteroid",
                         "sol": "Solar system", "gw": "gravitational waves", "dwM": "M-dwarfs", "gcl": "galaxy cluster",
                         "cv": "cataclysmic variable", "cos": "cosmology", "nea": "near-earth asteroids",
                         "psr": "pulsar", "cl": "cluster", "r-gal": "radio-galaxy", "wd": "white dwarf", "*": "star",
                         "ism": "interstellar matter", "loc": "local neighbourhood (Gal: sun; exg: loc univ)",
                         "yso": "young stellar object", "lss": "lss and distance measurements", "dw": "dwarf",
                         "V*": "variable star", "pn": "planetary nebulae", "WR*": "Wolf-Rayet star", "stb": "starburst",
                         "agn": "active galactic nucleus", "gl clu": "galaxy cluster", "tde": "tidal disruption event",
                         "*cl": "cluster of stars", "he*": "hydrogen deficient star", "uran": "Uranus",
                         "igm": "intergalactic medium", "dw*": "dwarf binary"}

    return all_science_types


def science_subject_and_explanation(science_type):
    subject = ""
    explanation = ""

    if pd.isnull(science_type) or "--" in science_type:
        subject = None
        explanation = None

    if not pd.isnull(science_type) and science_type.count("-") == 2 and "--" not in str(science_type):
        subject = "-".join(science_type.split("-")[1:])
        explanation = science_types()[subject.strip()]

    if not pd.isnull(science_type) and science_type.count("-") == 1 and "--" not in str(science_type):
        subject = (science_type.split("-")[-1]).strip()
        explanation = science_types()[subject]

    if not pd.isnull(science_type) and "-" not in science_type:
        subject = science_type
        explanation = science_types()[subject]

    return {"subject": subject, "explanation": explanation}


filepath = "/home/lonwabolap/Downloads/SALT publication statistics.xlsx"


def finding_column_for_flags():
    """
    We get the last column with data and two on top so as to account for the index
    since we count as in an array with the spreadsheet
    :return: The column number to add our flags for problems with proposals and publications
    """
    df = pd.read_excel(filepath, "Sheet1")
    columns_in_spreadsheet = df.columns
    num = [x for x in columns_in_spreadsheet if "Unnamed" not in x]
    column_to_add_flags_to = num.index(num[-1]) + 2
    return column_to_add_flags_to


book = openpyxl.load_workbook(filepath)


def add_value_to_cell(workbook_row, column, value, column_name):
    """
    We add a value to cell
    :param workbook_row: The row we need to add a value to
    :param column: The column to add a value to
    :param value: The value to be added in spreadsheet
    :param column_name: Name of our new column to add data to, we use "Flag"
    :return:
    """
    for worksheet in book.worksheets:
        for index, row in enumerate(worksheet.rows, start=1):
            worksheet.cell(row=1, column=column).value = column_name
            worksheet.cell(row=workbook_row, column=column).value = value
    return None


def start_of_discarded_papers():
    """
    We find the row which contains publications which have been discarded
    :return:
    """
    sheet = book.active
    worksheet = book["Sheet1"]
    discarded_papers = 0
    for row in sheet.iter_rows(min_row=1, max_col=worksheet.max_column, max_row=worksheet.max_row):
        for cell in row:
            if cell.value is not None and cell.value == "discarded papers:":
                discarded_papers = row[0].row
    return discarded_papers


def find_proposals_not_needed(grey_proposals):
    sheet = book.active
    worksheet = book["Sheet1"]
    arr = []
    for row in sheet.iter_rows(min_row=1, max_col=worksheet.max_column, max_row=worksheet.max_row):
        for cell in row:
            if cell.font.color is not None and cell.font.color.rgb in grey_proposals and row[0].row < \
                    start_of_discarded_papers():
                arr.append(row[0].row)
    return list(dict.fromkeys(arr))


def find_column_name():
    sheet = book.active
    worksheet = book["Sheet1"]
    columns = {}
    count = 1
    for col in sheet.iter_cols(1, worksheet.max_column):
        if col[0].value is not None:
            columns[count] = col[0].value
            count += 1
    return columns


def find_number_of_phd(proposal_code, df, value):
    cond1 = df["student project"] == "PhD,MSc"
    cond2 = df["student project"] == "PhD,PhD"
    row_number = df[cond1 | cond2].index[0] + 2
    number_of_students = 0

    if pd.isnull(proposal_code):
        number_of_students = 0
    if not pd.isnull(proposal_code):
        number_of_students = student_project_phd_numbers(proposal_code, 1)

    if str(value).strip() == "PhD,MSc".strip() or str(value).strip() == "PhD,PhD".strip():
        raise ValueError("The value for student project on row {} is inconsistent".format(row_number))

    return number_of_students


def find_number_of_msc(proposal_code, df, value):
    cond1 = df["student project"] == "PhD,MSc"
    cond2 = df["student project"] == "PhD,PhD"
    row_number = df[cond1 | cond2].index[0] + 2
    number_of_students = 0

    if pd.isnull(proposal_code):
        number_of_students = 0

    if not pd.isnull(proposal_code):
        number_of_students = student_project_msc_numbers(proposal_code, 2)

    if str(value).strip() == "PhD,MSc".strip() or str(value).strip() == "PhD,PhD".strip():
        raise ValueError("The value for student project on "
                         "row {} is inconsistent. Please correct it".format(row_number))
    return number_of_students


def find_proposal_semester(semester_on_spreadsheet, proposal_code):
    obj = {}

    if proposal_code is None:
        pass

    if semester_on_spreadsheet is not None and "." in str(semester_on_spreadsheet):
        year = str(semester_on_spreadsheet).split(".")[0]
        semester = str(semester_on_spreadsheet).split(".")[1]
        obj = {"year": year, "semester": semester}

    try:
        if semester_on_spreadsheet is None and proposal_code is not None:
            obj = {"year": semester_and_year_sdb(proposal_code)["year"],
                   "semester": semester_and_year_sdb(proposal_code)["semester"]}
    except IndexError:
        pass
    return obj


def create_dataframe(min_row, max_col, max_row):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb.active

    # This is for grey in rgb form (openpyxl)
    grey = ["FFB7B7B7", "FF999999"]
    # This is violet in rgb form (openpyxl)
    violet = ['FFA64D79']
    # This is green in rgb form (openpyxl)
    green = ["FF6AA84F"]
    # This is brown in rgb form (openpyxl)
    brown = ["FFB45F06", "FF783F04"]
    # red = ["FFFF0000"]
    # yellow for yellow background in rgb (openpyxl)
    yellow_fill = ["FFFFF2CC"]

    # red_indexes = []
    brown_indexes = []
    violet_indexes = []
    green_indexes = []
    discarded_papers = []
    yellow_indexes = []

    # We loop through the rows and columns of the spreadsheet and find the different color mentioned in the legend
    # and we create an array of all the rows which match the required colors for each
    # color(red, green, brown, yellow, violet and grey)
    for row in sheet.iter_rows(min_row=min_row, max_col=max_col, max_row=max_row):
        for cell in row:
            # We check if background color of cell is yellow and we add the row affected to the yellow_indexes array
            if cell.fill.fgColor.rgb in yellow_fill:
                yellow_indexes.append(row[0].row)

            # We get the row numbers in which discarded papers are in and append them to the discarded_papers array
            if cell.font.color is not None and cell.font.color.rgb in grey and int(row[0].row) > \
                    start_of_discarded_papers():
                discarded_papers.append(row[0].row)

            # We get the row numbers in which we have violet as font color and append to the violet_indexes array
            if cell.font.color is not None and cell.font.color.rgb in violet:
                violet_indexes.append(row[0].row)

            # We get the row numbers in which we have brown as font color and append to the brown_indexes array
            if cell.font.color is not None and cell.font.color.rgb in brown:
                brown_indexes.append(row[0].row)

            # We use the function find_column_name to get the name of column in which we have a cell that has a font
            # color of green and we create an object of that column row and the row number
            # Example: {"Proposal code(s)": 25} which says row 25 of column Proposal code(s) is green
            if cell.font.color is not None and cell.font.color.rgb in green:
                green_indexes.append({find_column_name()[cell.column]: row[0].row})

    # We then remove the last row numbers in each array so as not to include the colors in the rows of the legend
    violet_indexes.pop()
    brown_indexes.pop()
    discarded_papers.pop()
    green_indexes.pop()
    index_for_yellow = yellow_indexes[:-3]

    # 0 is for brown proposals
    for i in list(dict.fromkeys(brown_indexes)):
        add_value_to_cell(i, finding_column_for_flags(), 0, "Flag")

    # 1 is for violet proposals
    for i in list(dict.fromkeys(violet_indexes)):
        add_value_to_cell(i, finding_column_for_flags(), 1, "Flag")

    # 2 is for discarded publication papers
    for i in discarded_papers:
        add_value_to_cell(i, finding_column_for_flags(), 2, "Flag")

    # 3 is for proposals greyed out
    for i in find_proposals_not_needed(grey):
        add_value_to_cell(i, finding_column_for_flags(), 3, "Flag")

    # 4 is for yellow background indicating missing information
    for i in list(dict.fromkeys(index_for_yellow)):
        add_value_to_cell(i, finding_column_for_flags(), 4, "Flag")

    # This is to specify for the green font color, we check the cell(s) with green color and write a comment similar to
    # "Proposal code(s) not mentioned in paper, inferred. Also when we have a combination of colors such as
    # brown and green in a proposal or violet and green."

    my_dict = {}
    for i in green_indexes:
        for key, value in i.items():
            if value not in my_dict:
                my_dict[value] = []
            my_dict[value].append(key)

    for k, v in my_dict.items():
        add_value_to_cell(k, finding_column_for_flags(), "{} not mentioned in paper, inferred".format(",".join(v)),
                          "Flag")
        if k in violet_indexes:
            add_value_to_cell(k, finding_column_for_flags(),
                              "1, {} not mentioned in paper, inferred".format(",".join(v)),
                              "Flag")

        if k in brown_indexes:
            add_value_to_cell(k, finding_column_for_flags(),
                              "0, {} not mentioned in paper, inferred".format(",".join(v)),
                              "Flag")

    return wb


spreadsheet = pd.read_excel(filepath, "Sheet1")
workbook = create_dataframe(1, spreadsheet.shape[-1], spreadsheet.shape[0])
ws = workbook.active
data = ws.values
cols = next(data)[0:]


def proposal_code_error(proposal_code, df):

    if "and" in proposal_code or "," in proposal_code:
        indexing = df["Proposal code(s)"] == proposal_code
        raise ValueError("Incorrect proposal code '{}' on row {}".format(proposal_code, df[indexing].index[0] + 2))

    if " " in proposal_code:
        indexing = df["Proposal code(s)"] == proposal_code
        raise ValueError("Incorrect proposal code '{}' on row {}".format(proposal_code, df[indexing].index[0] + 2))

    if proposal_code.count("-") < 3:
        indexing = df["Proposal code(s)"] == proposal_code
        raise ValueError("Incorrect proposal code {} on row {}".format(proposal_code, df[indexing].index[0] + 2))


def proposal_issues(issue):
    obj = {3: "Proposal greyed out/ not counted towards publication"}

    if pd.isnull(issue) or issue not in obj:
        return "No issue found"

    if int(issue) in obj:
        return obj[int(issue)]


def publication_issues(issue):
    obj = {0: "data already presented in other SALT paper", 1: "Typo in paper", 2: "discarded papers",
           4: "todo/data missing"}

    number_in_flag = any(str.isdigit(i) for i in str(issue))

    if pd.isnull(issue) or isinstance(issue, int) not in obj:
        return "No issue found"

    if isinstance(issue, int) and issue in obj:
        return obj[issue]

    if "," in str(issue) and number_in_flag and "not mentioned" in str(issue):
        first_issue = re.findall("\d", issue)[0]
        second_issue = re.findall("\D", issue)
        return "combination of issues; {} and {}".format(obj[int(first_issue)], "".join(second_issue)[1:].strip())

    if not number_in_flag:
        return issue


def read_spreadsheet():
    # the spreadsheet read here should be the new and saved one after modifications(adding the flag)
    # filepath must be changed to point to that spreadsheet
    df = pd.read_excel(filepath, "Sheet1")

    arr = []
    for index, row in df.iterrows():

        if not pd.isnull(row["1st Author (status 1 November 2019)"]) and "violet" in row["1st Author (status 1 " \
                                                                                         "November 2019)"]:
            break

        if not (pd.isnull(row["1st Author (status 1 November 2019)"]) or
                row["1st Author (status 1 November 2019)"].strip() == "") and \
                "--" not in row["1st Author (status 1 November 2019)"] and \
                "discarded papers:" not in row["1st Author (status 1 November 2019)"]:
            arr.append({

                "Name": row["1st Author (status 1 November 2019)"],

                "ADS link": row["ADS link"],

                "Institute of 1st author": fixing_bad_column_return_none(row["Institute of 1st author"]),

                "Position of 1st author": fixing_bad_column_return_none(row["Position of 1st author"]),

                "Partnership of 1st author": fixing_bad_column_return_none(row["Partnership of 1st author"]),

                "Proposal code(s)": [{
                    "Proposal code": fixing_bad_column_return_none(row["Proposal code(s)"]),
                    "proposal semester": fixing_bad_column_return_none(row["Proposal semester"]),
                    "ToO": fixing_bad_column_return_none(row["ToO"]),
                    "PI": fixing_bad_column_return_none(row["PI"]),
                    "Partner(time allocated)": fixing_bad_column_return_none(row["Partner (time allocated)"]),
                    "Institutes (on proposal)": fixing_bad_column_return_none(row["Institutes (on proposal)"]),
                    "master's student": find_number_of_msc(row["Proposal code(s)"], df, row["student project"]),
                    "phd student": find_number_of_phd(row["Proposal code(s)"], df, row["student project"]),
                    "student project": fixing_bad_column_return_none(row["student project"]),
                    "Instrument(s)": fixing_bad_column_return_none(row["Instrument(s)"]),
                    "Instrument mode(s)": fixing_bad_column_return_none(row["Instrument mode(s)"]),
                    "observation date": fixing_bad_column_return_none(row["Dates obs (cf WM)"]),
                    "Priorities": fixing_bad_column_return_none(row["Priority (ies)"]),
                    "Total SALT time": fixing_bad_column_return_none(row["Total SALT time"]),
                    "Fraction of total time": fixing_bad_column_return_none(row["Fraction of total time [%]"]),
                    "Flag": fixing_bad_column_return_none(row["Flag"]),
                    "proposal issue": proposal_issues(row["Flag"])
                }],

                "Publication Paper": [{"Publication date": publication_date(row["Year"],
                                                                            row["Month (the ADS 'pub date')"], 15,
                                                                            df, row["ADS link"]),
                                       "Full author list": fixing_bad_column_return_none(row["Full author list"]),
                                       "Partners (on paper)": fixing_bad_column_return_none(
                                           row["Partners (on paper, excl 1st author)"]),
                                       "Institutes of partners (on paper, excl 1st author)":
                                           fixing_bad_column_return_none
                                           (row["Institutes of partners (on paper, excl 1st author)"]),
                                       "No of SA": fixing_bad_column_return_none(row["Num of SA on paper"]),
                                       "Comments": fixing_bad_column_return_none(row["Comments"]),
                                       "No of papers": fixing_bad_column_return_none(row["Number of papers"]),
                                       "Type of paper": fixing_bad_column_return_none(row["type of paper"]),
                                       "Type of Science": fixing_bad_column_return_none(row["type of science"]),
                                       "First Author Belonging": True,
                                       "Other Author Belonging": True if
                                       row["Partnership of 1st author"] == row["Partners (on paper, excl 1st author)"]
                                       else False,
                                       "publication issue": publication_issues(row["Flag"])
                                       }]

            })

        if pd.isnull(row["1st Author (status 1 November 2019)"]) or \
                row["1st Author (status 1 November 2019)"].strip() == "":
            arr[-1]["Proposal code(s)"].append({
                "Proposal code": fixing_bad_column_return_none(row["Proposal code(s)"]),
                "proposal semester": fixing_bad_column_return_none(row["Proposal semester"]),
                "ToO": fixing_bad_column_return_none(row["ToO"]),
                "PI": fixing_bad_column_return_none(row["PI"]),
                "Partner(time allocated)": fixing_bad_column_return_none(row["Partner (time allocated)"]),
                "Institutes (on proposal)": fixing_bad_column_return_none(row["Institutes (on proposal)"]),
                "master's student": find_number_of_msc(row["Proposal code(s)"], df, row["student project"]),
                "phd student": find_number_of_phd(row["Proposal code(s)"], df, row["student project"]),
                "student project": fixing_bad_column_return_none(row["student project"]),
                "Instrument(s)": fixing_bad_column_return_none(row["Instrument(s)"]),
                "Instrument mode(s)": fixing_bad_column_return_none(row["Instrument mode(s)"]),
                "Observation date": fixing_bad_column_return_none(row["Dates obs (cf WM)"]),
                "Priorities": fixing_bad_column_return_none(row["Priority (ies)"]),
                "Total SALT time": fixing_bad_column_return_none(row["Total SALT time"]),
                "Fraction of total time": fixing_bad_column_return_none(row["Fraction of total time [%]"]),
                "Flag": fixing_bad_column_return_none(row["Flag"]),
                "proposal issue": proposal_issues(row["Flag"])
            })

    return arr


def publication_information(publication_info):
    for information in publication_info:
        return information


def target_of_opportunity(spreadsheet_value):
    answer = ""
    if spreadsheet_value is None:
        answer = False

    if spreadsheet_value is not None and "yes" in spreadsheet_value:
        answer = True
    return answer


def is_row_red(row):
    red = 'FFFF0000'
    is_red = True
    for cell in row:
        if cell.value is not None and cell.font.color is not None and not cell.font.color.rgb == red:
            is_red = False
    return is_red


def is_partially_red(row):
    red = ['FFFF0000']
    is_red = False
    for cell in row:
        if cell.font.color is not None and cell.font.color.rgb in red:
            is_red = True
    return is_red


# same comment in read_spreadsheet here for filepath
dataframe = pd.read_excel(filepath, "Sheet1")

for values in read_spreadsheet():

    # insert into first author table
    if values["Name"]:
        salt_stats.insert_first_author(values["Name"], values["Position of 1st author"])

    if publication_information(values["Publication Paper"])["Type of Science"]:
        salt_stats.insert_science_category(publication_information(values["Publication Paper"])["Type of Science"])

    # insert into science subject table
    if publication_information(values["Publication Paper"])["Type of Science"]:
        salt_stats.insert_science_subject(science_subject_and_explanation(
            publication_information(values["Publication Paper"])["Type of Science"])["subject"],
                                          science_subject_and_explanation(publication_information(
                                              values["Publication Paper"])["Type of Science"])["explanation"],
                                          publication_information(values["Publication Paper"])["Type of Science"])

        # publication partner table should come after publication table

    # insert into Publication table
    salt_stats.insert_publication(values["Name"],
                                  publication_information(values["Publication Paper"])["Publication date"],
                                  values["ADS link"],
                                  science_subject_and_explanation(publication_information(
                                      values["Publication Paper"])["Type of Science"])["subject"],
                                  publication_information(values["Publication Paper"])["Type of Science"],
                                  publication_information(values["Publication Paper"])["Full author list"],
                                  publication_information(values["Publication Paper"])["No of SA"],
                                  publication_information(values["Publication Paper"])["Comments"]
                                  )

    # insert publication partnership of 1st author
    if values["Partnership of 1st author"]:
        salt_stats.insert_partner(values["Partnership of 1st author"])

        # insert publication partner
        salt_stats.insert_publication_partner(publication_information(values["Publication Paper"])["Publication date"],
                                              values["Name"],
                                              values["Partnership of 1st author"],
                                              publication_information(
                                                  values["Publication Paper"])["First Author Belonging"],
                                              publication_information(
                                                  values["Publication Paper"])["Other Author Belonging"],
                                              )
        # insert publication institute
        if values["Institution of 1st author"]:
            salt_stats.insert_publication_institute(
                publication_information(values["Publication Paper"])["Publication date"],
                values["Institute of 1st author"],
                values["Name"],
                publication_information(values["Publication Paper"])["First Author Belonging"],
                publication_information(values["Publication Paper"])["Other Author Belonging"])

    # insert proposal
    for value in values["Proposal code(s)"]:
        try:
            if value["Proposal code"]:
                salt_stats.insert_proposal(value["Proposal code"],
                                           proposal_investigator(value["Proposal code"])["ProposalInvestigator"],
                                           institutes(value["Proposal code"])["target of opportunity"],
                                           institutes(value["Proposal code"])["institutes"])
        except UnicodeEncodeError and TypeError:
            pass

    # insert student projects
    for value in values["Proposal code(s)"]:
        if value["Proposal code"]:
            salt_stats.insert_student_project(value["Proposal code"],
                                              value["master's student"],
                                              value["phd student"]
                                              )

    # insert time allocating partner
    for value in values["Proposal code(s)"]:
        if value["Proposal code"] and value["Partnership of 1st author"]:
            salt_stats.insert_time_allocating_partner(value["Proposal code"], value["Partnership of 1st author"])

    # insert instrument
    for value in values["Proposal code(s)"]:
        if value["Proposal code"] and value["Instrument(s)"]:
            salt_stats.insert_instrument(value["Instrument(s)"])

    # insert instrument mode
    for value in values["Proposal code(s)"]:
        if value["Proposal code"] and value["Instrument mode(s)"] and value["Instrument(s)"]:
            salt_stats.insert_instrument_mode(value["Instrument(s)"],
                                              value["Instrument mode(s)"])

    # insert semester
    for value in values["Proposal code(s)"]:
        if value["Proposal code"] and value["proposal semester"]:
            salt_stats.insert_semester(find_proposal_semester(value["proposal semester"],
                                                              value["Proposal code"]).get("year"),
                                       find_proposal_semester(value["proposal semester"],
                                                              value["Proposal code"]).get("semester")
                                       )

    # insert proposal instrument use
    salt_stats.insert_proposal_instrument_use(publication_information(
        values["Publication Paper"])["Publication date"],
                                              values["Name"],
                                              publication_information(values["Proposal code(s)"])["Proposal code"],
                                              publication_information(values["Proposal code(s)"])["Instrument mode(s)"],
                                              find_proposal_semester(publication_information(
                                                  values["Proposal code(s)"])["proposal semester"],
                                                                     publication_information(
                                                                         values["Proposal code(s)"])[
                                                                         "Proposal code"]).get("year"),
                                              find_proposal_semester(publication_information(
                                                  values["Proposal code(s)"])["proposal semester"],
                                                                     publication_information(
                                                                         values["Proposal code(s)"])[
                                                                         "Proposal code"]).get("semester"),
                                              publication_information(
                                                  values["Proposal code(s)"])["observation date"],
                                              publication_information(values["Proposal code(s)"])["Priorities"],
                                              publication_information(
                                                  values["Proposal code(s)"])["Total SALT time"],
                                              publication_information(
                                                  values["Proposal code(s)"])["Fraction of total time"])
    # issues with proposals
    for value in values["Proposal code(s)"]:
        if value["Proposal code"]:
            salt_stats.insert_actual_issues_with_proposals(proposal_issues(value["proposal issue"]))

    # issues for all publications
    salt_stats.insert_actual_issues_with_publications(publication_issues(publication_information(
        values["Proposal code(s)"])["Flag"]))

    # Then we insert to the ProposalIssues tables
    for value in values["Proposal code(s)"]:
        if value["Proposal code"]:
            salt_stats.insert_proposal_issues(value["Proposal code"], value["proposal issue"])

    # Then we insert to the PublicationIssues table
    salt_stats.insert_publication_issues(publication_information(
        values["Publication Paper"])["Publication date"], values["Name"],
                                         publication_information(values["Publication Paper"])["publication issue"])
